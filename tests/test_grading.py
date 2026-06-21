"""Grading tests for the GDPval template (v6).

Focus on the parts that carry real scoring risk — the blended deterministic +
LLM-judge grade, the no-key degrade path, and the fabrication hard cap. The judge
is monkeypatched so the whole suite runs offline (no HUD gateway / key).
"""

import importlib.util
import json
import shutil
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import deliverable_io  # noqa: E402
import native_grading  # noqa: E402
from hud.graders import SubScore  # noqa: E402

ACCT = "acct-afc-audit-sampling"


def _rubric(slug: str) -> dict:
    return json.loads((ROOT / "tasks" / slug / "_hidden" / "rubric.json").read_text())


def _load_grader(slug: str):
    spec = importlib.util.spec_from_file_location(f"_t_grader_{slug.replace('-', '_')}", ROOT / "tasks" / slug / "grader.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _patch_judge(monkeypatch, *, fabrication_value: float, quality_value: float = 1.0):
    """Replace the live LLM judge with a deterministic stub keyed on subscore name."""

    async def _fake(cls, *, name=None, weight=0.0, **kwargs):
        value = fabrication_value if name == "fabrication_guard" else quality_value
        return SubScore(name=name, weight=weight, value=value)

    monkeypatch.setattr(native_grading, "judge_available", lambda: True)
    monkeypatch.setattr(native_grading.LLMJudgeGrader, "grade", classmethod(_fake))


async def test_det_only_when_no_key(monkeypatch):
    """With no judge key, the deterministic subscore keeps its real weight (must not
    renormalize to 1.0) and the judge component scores 0 at its weight."""
    monkeypatch.setattr(native_grading, "judge_available", lambda: False)
    out = await native_grading.blend_with_native_judge(
        det_score=0.8, det_weight=0.45, det_detail={"probe": 1},
        key=_rubric(ACCT), axis_weights={"factual_accuracy": 1.0},
        submitted="deliverable text", extra_instruction="",
    )
    assert out["info"]["status"] == "det_only:no_hud_api_key"
    # reward = det_weight * det_score = 0.45 * 0.8 = 0.36 (judge contributes 0).
    assert out["reward"] == pytest.approx(0.36, abs=1e-6)
    subs = {s["name"]: s for s in out["info"]["subscores"]}
    assert subs["deterministic"]["weight"] == pytest.approx(0.45)
    assert subs["llm_judge"]["weight"] == pytest.approx(0.55)
    assert subs["llm_judge"]["value"] == 0.0


async def test_fabrication_hard_caps_reward(monkeypatch):
    """A deliverable the fabrication guard flags (value < 0.5) is capped at
    FABRICATION_CAP even when the deterministic + quality blend scores higher."""
    _patch_judge(monkeypatch, fabrication_value=0.0, quality_value=1.0)
    out = await native_grading.blend_with_native_judge(
        det_score=1.0, det_weight=0.45, det_detail={},
        key=_rubric(ACCT), axis_weights={"factual_accuracy": 1.0},
        submitted="text", extra_instruction="",
    )
    assert out["info"]["fabrication_capped"] is True
    assert out["reward"] == pytest.approx(native_grading.FABRICATION_CAP)


async def test_no_fabrication_no_cap(monkeypatch):
    """Clean deliverable: det 1.0 (w=0.45) + quality 1.0 (w=0.55) -> reward 1.0, uncapped."""
    _patch_judge(monkeypatch, fabrication_value=1.0, quality_value=1.0)
    out = await native_grading.blend_with_native_judge(
        det_score=1.0, det_weight=0.45, det_detail={},
        key=_rubric(ACCT), axis_weights={"factual_accuracy": 1.0},
        submitted="text", extra_instruction="",
    )
    assert out["info"]["fabrication_capped"] is False
    assert out["reward"] == pytest.approx(1.0)


def test_grading_result_fits_control_channel_frame():
    """The result is framed as one <64KB JSON line over the control channel, but the
    SDK judge echoes the full answer into each subscore's _parameters — so a large
    deliverable blows the frame unless native_grading strips it. Reverting the strip
    makes this fail (the acct payload is ~220KB)."""
    from hud.graders.combine import _combine_subscores

    big = "X" * 110_000  # roughly submitted[:50k] + reference_context[:60k]
    meta = {"criteria": {"axis": {"verdict": "MET"}}, "model": "m",
            "_parameters": {"answer": big, "criteria": ["..."], "question": "Grade."}}
    subs = [
        SubScore(name="deterministic", value=0.5, weight=0.45, metadata={"checks": {}}),
        SubScore(name="llm_judge", value=0.8, weight=0.55, metadata=dict(meta)),
        SubScore(name="fabrication_guard", value=1.0, weight=0.0, metadata=dict(meta)),
    ]
    out = native_grading._as_dict(_combine_subscores(subs), status="ok", deterministic={"checks": {}})
    frame = json.dumps(out).encode()
    assert len(frame) < 65_536, f"grading frame {len(frame)} bytes exceeds the 64KB wire limit"


def test_load_grader_resolves_siblings_without_repo_on_path(monkeypatch):
    """Regression: under a served env the repo dir isn't on sys.path at grade time, so
    `env._load_grader` must re-assert it for the grader's sibling imports to resolve.
    Reverting that fix makes this raise ModuleNotFoundError."""
    import env

    repo = str(env.APP_ROOT)
    pruned = [p for p in sys.path if p not in ("", ".") and Path(p).resolve() != Path(repo)]
    monkeypatch.setattr(sys, "path", pruned)
    for name in ("deliverable_io", "native_grading"):
        monkeypatch.delitem(sys.modules, name, raising=False)
    assert repo not in sys.path, "test setup: repo dir must be off sys.path"

    src = (env.APP_ROOT / "tasks" / ACCT / "grader.py").read_text()
    mod = env._load_grader(ACCT, src)
    assert hasattr(mod, "grade"), "grader must load with its grade() entrypoint"


async def test_real_acct_deterministic_axis(tmp_path, monkeypatch):
    """Deterministic axis end to end: build a sample workbook from the real GDPval
    population, grade with no key, and assert a sensible det-only reward."""
    from openpyxl import Workbook, load_workbook

    monkeypatch.setattr(native_grading, "judge_available", lambda: False)
    ws_ref = tmp_path / "reference_files"
    shutil.copytree(ROOT / "tasks" / ACCT / "reference_files", ws_ref)

    pop = load_workbook(str(ws_ref / "Population v2.xlsx"), read_only=True, data_only=True)
    rows = list(pop.worksheets[0].iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    no_i = header.index("No")
    sample_ids = [r[no_i] for r in rows[1:40] if r[no_i] is not None][:12]

    wb = Workbook()
    calc = wb.active
    calc.title = "Sample Size Calculation"
    calc.append(["Sample size at 90% confidence, 10% tolerable error"])
    calc.append(["variance computed quarter-on-quarter below"])
    sample = wb.create_sheet("Sample")
    sample.append(["No", *header[1:]])
    src = {str(r[no_i]): r for r in rows[1:] if r[no_i] is not None}
    for rid in sample_ids:
        sample.append(list(src[str(rid)]))
    deliv = tmp_path / "deliverable" / "Sample.xlsx"
    deliv.parent.mkdir(parents=True)
    wb.save(str(deliv))

    out = await _load_grader(ACCT).grade(tmp_path, deliv, _rubric(ACCT))
    assert out["info"]["status"] == "det_only:no_hud_api_key"
    # det-only, so reward <= det_weight; a real sample should clear zero.
    assert 0.0 < out["reward"] <= 0.45 + 1e-9
    det = out["info"]["deterministic"]["checks"]
    assert det["parseable_xlsx"] == 1.0
    assert det["is_a_sample_not_whole_population"] == 1.0
